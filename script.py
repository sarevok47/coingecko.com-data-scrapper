from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font

def clean_digit(x):
    if not x:
        return None
    cleaned = x.replace('$', '').replace(',', '')
    return float(cleaned)

with (sync_playwright() as p):
    browser = p.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://www.coingecko.com/", timeout=60000)

    wb = Workbook()
    ws = wb.active
    ws.title = "coingecko coins"

    headers = ["Name", "Price", "Market Cap"]
    ws.append(headers)
    ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)

    ws.column_dimensions['A'].width = ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width  = 40

    while True:
      page.wait_for_selector('tr.hover\\:tw-bg-gray-50')

      for el in page.query_selector_all('tr[class="hover:tw-bg-gray-50 tw-bg-white dark:tw-bg-moon-900 hover:dark:tw-bg-moon-800 tw-text-sm"]'):
        coin_el =  el.query_selector("td:nth-of-type(3) div.tw-text-gray-700")
        coin_name = coin_el.text_content().strip().split('\n')[0].strip().split()[0]
        price_el = el.query_selector('td:nth-child(5) span[data-price-target="price"]')
        price = clean_digit(price_el.text_content().strip()) if price_el else ""
        market_cap_el =  el.query_selector('td:nth-child(10) span[data-price-target="price"]')
        market_cap = clean_digit(market_cap_el.text_content().strip()) if market_cap_el else ""
        ws.append([coin_name, price, market_cap])

      nav_el = page.query_selector(
          'nav.tailwind-reset.pagy_nav.tw-inline-flex.tw-items-center.tw-space-x-1.tw-rounded-lg.gecko-pagination-nav')

      next_link = nav_el.query_selector_all("span a[href]")[-1]
      if next_link:
          next_link.click()
      else:
          break
    wb.save("sample_data.xlsx")

